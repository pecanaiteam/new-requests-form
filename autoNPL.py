import subprocess
import time
import requests
import re
import os
import base64
from flask import Flask, request, jsonify, make_response
from flask_cors import CORS
from datetime import datetime
from fpdf import FPDF
from openpyxl import load_workbook, Workbook

# === INIT FLASK APP ===
app = Flask(__name__)
CORS(app, supports_credentials=True)

# === AUTO DEPLOY SECTION ===
RENDER_SERVICE_ID = "srv-d0r3fbfdiees73blq330"
RENDER_API_KEY = "rnd_PVTkldXhRNq8DwWypV9FzBihMVjd"

def start_ngrok():
    ngrok_proc = subprocess.Popen(["ngrok", "http", "5002"], stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)
    time.sleep(5)
    response = requests.get("http://localhost:4040/api/tunnels")
    public_url = response.json()["tunnels"][0]["public_url"]
    print(f"[+] Ngrok URL: {public_url}")
    return ngrok_proc, public_url

def update_index_html(public_url):
    with open("index.html", "r", encoding="utf-8") as file:
        html = file.read()
    updated = re.sub(r"https?://[a-z0-9\-]+\.ngrok-free\.app|http://localhost:5002", public_url, html)
    if html != updated:
        with open("index.html", "w", encoding="utf-8") as file:
            file.write(updated)
        print("[+] index.html updated with new ngrok URL")
        return True
    else:
        print("[~] No changes needed — URL already current.")
        return False

def commit_and_push_changes():
    subprocess.run(["git", "add", "index.html"], check=True)
    subprocess.run(["git", "commit", "-m", "Auto-update ngrok URL"], check=False)
    subprocess.run(["git", "push", "--set-upstream", "origin", "main"], check=False)
    print("[+] Pushed to GitHub")

def trigger_render_deploy():
    headers = {
        "Authorization": f"Bearer {RENDER_API_KEY}",
        "Accept": "application/json",
        "Content-Type": "application/json"
    }
    response = requests.post(
        f"https://api.render.com/v1/services/{RENDER_SERVICE_ID}/deploys",
        headers=headers
    )
    if response.status_code == 201:
        print("[🚀] Render deployment triggered successfully.")
    else:
        print(f"[!] Failed to trigger deployment: {response.text}")

def auto_deploy():
    ngrok_proc, new_url = start_ngrok()
    try:
        changed = update_index_html(new_url)
        if changed:
            commit_and_push_changes()
        else:
            print("[~] Skipped commit and push.")
        trigger_render_deploy()
    finally:
        ngrok_proc.terminate()

# === FORM PROCESSING ===
SAVE_FOLDER = "submissions"
EXCEL_PATH = os.path.join(SAVE_FOLDER, "submissions.xlsx")

os.makedirs(SAVE_FOLDER, exist_ok=True)

if not os.path.exists(EXCEL_PATH):
    wb = Workbook()
    ws = wb.active
    ws.append(["Timestamp", "Name", "DOB", "Phone", "Email", "Insurance", "Reason"])
    wb.save(EXCEL_PATH)

class PDF(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'New Patient Form', ln=True, align='C')
        self.ln(10)

    def add_patient_info(self, data):
        self.set_font('Arial', '', 12)
        for label, value in data.items():
            if label != 'signature':
                self.cell(0, 10, f"{label}: {value}", ln=True)

    def add_signature(self, signature_path):
        self.ln(10)
        self.cell(0, 10, "Signature:", ln=True)
        self.image(signature_path, w=60)

@app.route("/submit", methods=["POST", "OPTIONS"])
def submit():
    if request.method == "OPTIONS":
        response = make_response()
        response.headers["Access-Control-Allow-Origin"] = "https://new-patient-form.onrender.com"
        response.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type"
        return response

    try:
        data = request.get_json()
        print("[DEBUG] Received:", data)

        # (same as before)
        timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
        signature_data = data.get("signature")
        signature_path = os.path.join(SAVE_FOLDER, f"signature_{timestamp}.png")
        pdf_path = os.path.join(SAVE_FOLDER, f"form_{timestamp}.pdf")

        with open(signature_path, "wb") as sig_file:
            sig_file.write(base64.b64decode(signature_data.split(",")[1]))

        pdf = PDF()
        pdf.add_page()
        pdf.add_patient_info({
            "Name": data.get("name"),
            "DOB": data.get("dob"),
            "Phone": data.get("phone"),
            "Email": data.get("email"),
            "Insurance": data.get("insurance"),
            "Reason": data.get("reason"),
            "Timestamp": timestamp
        })
        pdf.add_signature(signature_path)
        pdf.output(pdf_path)

        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        ws.append([
            timestamp,
            data.get("name"),
            data.get("dob"),
            data.get("phone"),
            data.get("email"),
            data.get("insurance"),
            data.get("reason")
        ])
        wb.save(EXCEL_PATH)

        response = jsonify({"status": "success"})
        response.headers["Access-Control-Allow-Origin"] = "https://new-patient-form.onrender.com"
        return response

    except Exception as e:
        response = jsonify({"status": "error", "message": str(e)})
        response.headers["Access-Control-Allow-Origin"] = "https://new-patient-form.onrender.com"
        return response, 500


# === START APP ===
if __name__ == "__main__":
    auto_deploy()
    app.run(host="0.0.0.0", port=5002)
