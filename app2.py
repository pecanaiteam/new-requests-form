import os
import re
from pyngrok import ngrok
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename
from github import Github, GithubException
from datetime import datetime

# 🔐 User login credentials
VALID_USERS = {
    "admin": "password123",
    "user1": "letmein"
}

# Start an http‐only ngrok tunnel on port 5000
ngrok.kill()
tunnel = ngrok.connect(5000)
public_url = tunnel.public_url
print("ngrok public url:", public_url)

# Inject the new ngrok URL into index.html
INDEX_PATH = "index.html"
with open(INDEX_PATH, "r", encoding="utf-8") as f:
    html = f.read()
new_action = f'action="{public_url}/submit"'
updated_html = re.sub(r'action="https?://[^"]+/submit"', new_action, html)
if updated_html != html:
    with open(INDEX_PATH, "w", encoding="utf-8") as f:
        f.write(updated_html)
    print("index.html updated with new ngrok URL")

# Push index.html and this file to GitHub
token = os.getenv("GITHUB_TOKEN")
repo_name = os.getenv("GITHUB_REPO")
if token and repo_name:
    gh = Github(token)
    try:
        repo = gh.get_repo(repo_name)
        for fname in (INDEX_PATH, os.path.basename(__file__)):
            with open(fname, "r", encoding="utf-8") as f:
                content = f.read()
            try:
                contents = repo.get_contents(fname, ref="main")
                repo.update_file(
                    contents.path,
                    f"Auto-update {fname}",
                    content,
                    contents.sha,
                    branch="main"
                )
                print(f"updated {fname} on github")
            except GithubException:
                repo.create_file(
                    fname,
                    f"Add {fname}",
                    content,
                    branch="main"
                )
                print(f"created {fname} on github")
    except Exception as e:
        print("github push error:", e)
else:
    print("set GITHUB_TOKEN and GITHUB_REPO env vars to enable github push")

# Flask and Excel setup
app = Flask(__name__)
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "submissions.xlsx")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Create Excel with headers if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Requestor Name", "Dealer Name", "Email", "Phone",
        "Feature 1 Priority", "Feature 1 Description", "Feature 1 Severity", "Feature 1 Attachment",
        "Feature 2 Priority", "Feature 2 Description", "Feature 2 Severity", "Feature 2 Attachment",
        "Feature 3 Priority", "Feature 3 Description", "Feature 3 Severity", "Feature 3 Attachment"
    ])
    wb.save(EXCEL_FILE)

@app.route("/", methods=["GET"])
def index():
    return send_from_directory(BASE_DIR, INDEX_PATH)

# ✅ New Login route
@app.route("/login", methods=["POST"])
def login():
    try:
        data = request.get_json()
        username = data.get("username")
        password = data.get("password")

        if VALID_USERS.get(username) == password:
            return jsonify({"success": True}), 200
        else:
            return jsonify({"error": "Invalid credentials"}), 401
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/submit", methods=["POST"])
def submit():
    try:
        form = request.form
        files = request.files
        print("Received form:", form)
        print("Received files:", files)

        row = [
            form.get("requestor_name", ""),
            form.get("dealer_name", ""),
            form.get("email", ""),
            form.get("phone", "")
        ]

        priority_map = {"1": "Urgent", "2": "Normal", "3": "Optional"}
        severity_map = {
            "1": "Cannot Operate/Sell without",
            "2": "Important but Workable",
            "3": "Nice to Have"
        }

        for i in range(1, 4):
            priority_raw = form.get(f"priority_{i}", "")
            priority = priority_map.get(priority_raw, priority_raw)
            desc = form.get(f"feature_description_{i}", "")
            severity_raw = form.get(f"severity_{i}", "")
            sev = severity_map.get(severity_raw, severity_raw)

            upload = files.get(f"attachment_{i}")
            fname = ""
            if upload and upload.filename:
                raw_dealer = form.get("dealer_name", "dealer")
                dealer = secure_filename(raw_dealer.replace(" ", "_"))
                date_str = datetime.now().strftime("%Y%m%d")
                _, ext = os.path.splitext(upload.filename)
                fname = secure_filename(f"{dealer}_feature{i}_{date_str}{ext}")
                upload_path = os.path.join(UPLOAD_FOLDER, fname)
                upload.save(upload_path)
                print(f"Saved attachment {i} as:", upload_path)

            row.extend([priority, desc, sev, fname])

        print("Appending row:", row)
        print("Using Excel path:", EXCEL_FILE)
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append(row)
        wb.save(EXCEL_FILE)
        print("Excel saved.")

        return jsonify({"status": "success"})
    except Exception as e:
        print("Error occurred:", str(e))
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
