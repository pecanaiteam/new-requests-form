import os
import re
from pyngrok import ngrok
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename
from github import Github, GithubException
from datetime import datetime


# ---------- ngrok ----------
ngrok.kill()
tunnel = ngrok.connect(5000)
public_url = tunnel.public_url
print("ngrok public url:", public_url)

# ---------- Inject URLs into index.html ----------
INDEX_PATH = "index.html"
with open(INDEX_PATH, "r", encoding="utf-8") as f:
    html = f.read()

# Update form action
new_action = f'action="{public_url}/submit"'
updated_html = re.sub(r'action="https?://[^"]+/submit"', new_action, html)

# Update BACKEND_URL for votes
backend_line = f'const BACKEND_URL = "{public_url}/feature-vote";'
if 'const BACKEND_URL' in updated_html:
    updated_html = re.sub(r'const BACKEND_URL\s*=\s*".*?";', backend_line, updated_html)
else:
    # If not present (should be), append it near the end before </script>
    updated_html = updated_html.replace(
        "</script>",
        f"\n// injected by server\n{backend_line}\n</script>"
    )

if updated_html != html:
    with open(INDEX_PATH, "w", encoding="utf-8") as f:
        f.write(updated_html)
    print("index.html updated with new ngrok URLs")

# ---------- Optional GitHub push ----------
token = os.getenv("GITHUB_TOKEN")
repo_name = os.getenv("GITHUB_REPO")
if token and repo_name:
    gh = Github(token)
    try:
        repo = gh.get_repo(repo_name)
        for fname in (INDEX_PATH, os.path.basename(__file__)):
            with open(fname, "r", encoding="utf-8") as f:
                content = f.read()
            try:
                contents = repo.get_contents(fname, ref="main")
                repo.update_file(
                    contents.path,
                    f"Auto-update {fname}",
                    content,
                    contents.sha,
                    branch="main"
                )
                print(f"updated {fname} on github")
            except GithubException:
                repo.create_file(
                    fname,
                    f"Add {fname}",
                    content,
                    branch="main"
                )
                print(f"created {fname} on github")
    except Exception as e:
        print("github push error:", e)
else:
    print("set GITHUB_TOKEN and GITHUB_REPO env vars to enable github push")

# ---------- Flask / Excel setup ----------
app = Flask(__name__)
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "submissions.xlsx")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

REQUEST_SHEET = "Sheet"            # your default first sheet for submissions
VOTE_SHEET    = "FeatureVotes"     # new sheet for vote tallies

def ensure_workbook():
    """Ensure workbook exists with both sheets/headers."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = REQUEST_SHEET
        ws.append([
            "Requestor Name", "Dealer Name", "Email", "Phone",
            "Feature 1 Priority", "Feature 1 Description", "Feature 1 Severity", "Feature 1 Attachment",
            "Feature 2 Priority", "Feature 2 Description", "Feature 2 Severity", "Feature 2 Attachment",
            "Feature 3 Priority", "Feature 3 Description", "Feature 3 Severity", "Feature 3 Attachment"
        ])
        # create votes sheet
        vs = wb.create_sheet(VOTE_SHEET)
        vs.append(["Feature ID", "Summary", "Votes_No", "Votes_Nice", "Votes_Must", "Last Updated"])
        wb.save(EXCEL_FILE)
        return

    wb = load_workbook(EXCEL_FILE)
    if VOTE_SHEET not in wb.sheetnames:
        vs = wb.create_sheet(VOTE_SHEET)
        vs.append(["Feature ID", "Summary", "Votes_No", "Votes_Nice", "Votes_Must", "Last Updated"])
        wb.save(EXCEL_FILE)

ensure_workbook()

def open_book():
    return load_workbook(EXCEL_FILE)

def get_vote_sheet(wb):
    return wb[VOTE_SHEET]

def find_vote_row(ws, feature_id):
    """Return (row_index) where Feature ID matches, or None."""
    for r in range(2, ws.max_row + 1):
        if (ws.cell(r, 1).value or "") == feature_id:
            return r
    return None

def read_vote_row(ws, r):
    """Read counts safely (ints)."""
    def _to_int(v): 
        try: return int(v or 0)
        except: return 0
    return {
        "no":   _to_int(ws.cell(r, 3).value),
        "nice": _to_int(ws.cell(r, 4).value),
        "must": _to_int(ws.cell(r, 5).value),
    }

def write_vote_row(ws, r, counts, summary=None):
    ws.cell(r, 3, max(0, counts["no"]))
    ws.cell(r, 4, max(0, counts["nice"]))
    ws.cell(r, 5, max(0, counts["must"]))
    if summary is not None:
        ws.cell(r, 2, summary)
    ws.cell(r, 6, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

def bump(ws, feature_id, summary, inc_no=0, inc_nice=0, inc_must=0):
    """Increment counts for a feature (create row if needed)."""
    r = find_vote_row(ws, feature_id)
    if r is None:
        r = ws.max_row + 1
        ws.cell(r, 1, feature_id)
        ws.cell(r, 2, summary or "")
        ws.cell(r, 3, 0)
        ws.cell(r, 4, 0)
        ws.cell(r, 5, 0)
        ws.cell(r, 6, "")

    counts = read_vote_row(ws, r)
    counts["no"]   += inc_no
    counts["nice"] += inc_nice
    counts["must"] += inc_must
    write_vote_row(ws, r, counts, summary=summary)
    return counts

def adjust_for_prev(ws, feature_id, prev_choice):
    """If prev_choice provided, decrement that column by 1 (not below 0)."""
    if not prev_choice:
        return
    r = find_vote_row(ws, feature_id)
    if r is None:
        return
    counts = read_vote_row(ws, r)
    key = {"no":"no", "nice":"nice", "must":"must"}.get(prev_choice)
    if key and counts[key] > 0:
        counts[key] -= 1
        write_vote_row(ws, r, counts)

@app.route("/", methods=["GET"])
def index():
    return send_from_directory(BASE_DIR, INDEX_PATH)

# ---------- Existing form submission ----------
@app.route("/submit", methods=["POST"])
def submit():
    try:
        form = request.form
        files = request.files
        print("Received form:", form)
        print("Received files:", files)

        row = [
            form.get("requestor_name", ""),
            form.get("dealer_name", ""),
            form.get("email", ""),
            form.get("phone", "")
        ]

        # NOTE: Your mapping here looked reversed vs. the HTML, but leaving as-is per your code.
        priority_map = {"1": "Urgent", "2": "Normal", "3": "Optional"}
        severity_map = {
            "1": "Cannot Operate/Sell without",
            "2": "Important but Workable",
            "3": "Nice to Have"
        }

        for i in range(1, 4):
            priority_raw = form.get(f"priority_{i}", "")
            priority = priority_map.get(priority_raw, priority_raw)

            desc = form.get(f"feature_description_{i}", "")

            severity_raw = form.get(f"severity_{i}", "")
            sev = severity_map.get(severity_raw, severity_raw)

            upload = files.get(f"attachment_{i}")
            fname = ""
            if upload and upload.filename:
                raw_dealer = form.get("dealer_name", "dealer")
                dealer = secure_filename(raw_dealer.replace(" ", "_"))
                date_str = datetime.now().strftime("%Y%m%d")
                _, ext = os.path.splitext(upload.filename)
                fname = secure_filename(f"{dealer}_feature{i}_{date_str}{ext}")
                upload_path = os.path.join(UPLOAD_FOLDER, fname)
                upload.save(upload_path)
                print(f"Saved attachment {i} as:", upload_path)

            row.extend([priority, desc, sev, fname])

        print("Appending row:", row)
        print("Using Excel path:", EXCEL_FILE)
        wb = open_book()
        ws = wb[REQUEST_SHEET]
        ws.append(row)
        wb.save(EXCEL_FILE)
        print("Excel saved (request row).")

        return jsonify({"status": "success"})
    except Exception as e:
        print("Error occurred:", str(e))
        return jsonify({"status": "error", "message": str(e)}), 500

# ---------- NEW: feature vote endpoint ----------
@app.route("/feature-vote", methods=["POST"])
def feature_vote():
    """
    Accepts:
      - Bulk: { "votes": [ { "id": "...", "choice": "no|nice|must", "summary": "..." }, ... ] }
      - Single: { "id": "...", "choice": "no|nice|must", "prev_choice": "no|nice|must" (optional), "summary": "..." }
    Updates FeatureVotes sheet by incrementing the appropriate counters.
    Returns updated totals for touched features.
    """
    try:
        payload = request.get_json(silent=True) or {}
        wb = open_book()
        vs = get_vote_sheet(wb)

        touched = {}

        def apply_one(v):
            fid = v.get("id")
            choice = v.get("choice")
            prev_choice = v.get("prev_choice")
            summary = v.get("summary", "")  # optional; stored if provided
            if not fid or choice not in ("no", "nice", "must"):
                return

            # decrement previous choice if provided (for per-click mode)
            if prev_choice in ("no", "nice", "must"):
                adjust_for_prev(vs, fid, prev_choice)

            inc = {"no":0, "nice":0, "must":0}
            inc[choice] = 1
            counts = bump(vs, fid, summary, inc_no=inc["no"], inc_nice=inc["nice"], inc_must=inc["must"])
            touched[fid] = {
                "votes_no": counts["no"],
                "votes_nice": counts["nice"],
                "votes_must": counts["must"]
            }

        if "votes" in payload and isinstance(payload["votes"], list):
            for v in payload["votes"]:
                apply_one(v)
        else:
            apply_one(payload)

        wb.save(EXCEL_FILE)
        return jsonify({"status": "ok", "totals": touched})
    except Exception as e:
        print("feature-vote error:", e)
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == "__main__":
    ensure_workbook()
    app.run(host="0.0.0.0", port=5000)
